import streamlit as st
import pandas as pd
from io import StringIO, BytesIO
from thefuzz import fuzz
import re
import unicodedata
import openpyxl
from copy import copy
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.utils.cell import range_boundaries, get_column_letter
from openpyxl.styles import PatternFill
import logging

# --- CONFIGURACIÓN ---
logging.basicConfig(level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')
st.set_page_config(page_title="Control PSA V33.0", layout="wide", page_icon="🛡️")

# --- ESTILOS CSS ---
st.markdown("""
<style>
    /* 1. RESET Y FUENTES */
    * { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif; }
    .block-container { padding-top: 1rem; padding-bottom: 5rem; max-width: 1400px; }
    
    /* 2. BOTONES */
    div.stButton > button {
        border-radius: 6px; font-weight: 800 !important; font-size: 13px;
        height: 40px; border: none; width: 100%; transition: all 0.2s;
    }

    /* PRIMARIO (AMARILLO) */
    div.stButton > button[kind="primary"] {
        background-color: #F5A623 !important; color: #000 !important; border: 1px solid #F5A623 !important;
    }
    div.stButton > button[kind="primary"]:hover {
        background-color: #D48806 !important; transform: scale(1.02);
    }
    
    /* SECUNDARIO (ROJO) */
    div.stButton > button[kind="secondary"] {
        background-color: #2C1A1A !important; color: #FF5555 !important; border: 1px solid #7f1d1d !important;
    }
    div.stButton > button[kind="secondary"]:hover {
        background-color: #DC2626 !important; color: #FFF !important; border-color: #DC2626 !important;
    }

    /* 3. TARJETAS DETECTIVE */
    .conflict-container {
        background-color: #111218; border: 1px solid #333; border-radius: 6px;
        padding: 0; display: grid; grid-template-columns: 140px 1fr;
        align-items: center; height: 40px; overflow: hidden;
    }
    .c-badge {
        background-color: #1A1B25; color: #888; font-size: 10px; font-weight: 700;
        text-transform: uppercase; display: flex; align-items: center; justify-content: center; 
        height: 100%; border-right: 1px solid #333;
    }
    .c-name {
        color: #FFF; font-size: 13px; font-weight: 700; text-transform: uppercase;
        padding-left: 12px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
    }

    /* 4. TABLAS */
    .row-container {
        display: flex; align-items: center; height: 38px; border-bottom: 1px solid #222; margin-bottom: 2px;
    }
    .unified-text {
        font-size: 12px; font-weight: 600; color: #EEE; text-transform: uppercase;
        white-space: nowrap; overflow: hidden; text-overflow: ellipsis; display: block;
    }
    .name-ready { background-color: rgba(34, 197, 94, 0.2); color: #4ade80; border-radius: 4px; padding: 2px 8px; width: 100%; }

    /* UTILS */
    .stCode { font-family: sans-serif !important; }
    [data-testid="stCodeBlock"] button { color: #666 !important; }
    hr { margin: 15px 0 !important; border-color: #333 !important; opacity: 1; }
    .duplicate-alert { padding: 8px; background: rgba(245, 158, 11, 0.1); color: #F59E0B; border: 1px solid #F59E0B; font-size: 12px; border-radius: 4px; margin-bottom: 10px; }
    .stTextArea textarea, .stTextInput input { background-color: #0E0E10 !important; border: 1px solid #333 !important; color: #FFF !important; }
</style>
""", unsafe_allow_html=True)

st.title("🛡️ CONTROL DE PERSONAL V33")

# --- CONSTANTES ---
EQUIVALENCIAS = {
    "oficial ayudante": "OFICIAL AYUDANTE", "of ayte": "OFICIAL AYUDANTE", "of. ayte": "OFICIAL AYUDANTE", "ayte": "OFICIAL AYUDANTE",
    "oficial principal": "OFICIAL PRINCIPAL", "of ppal": "OFICIAL PRINCIPAL", "of. ppal": "OFICIAL PRINCIPAL", "ppal": "OFICIAL PRINCIPAL",
    "oficial mayor": "OFICIAL MAYOR", "of mayor": "OFICIAL MAYOR", "of. mayor": "OFICIAL MAYOR",
    "oficial jefe": "OFICIAL JEFE", "of jefe": "OFICIAL JEFE", "of. jefe": "OFICIAL JEFE",
    "subinspector": "SUBINSPECTOR", "sub inspector": "SUBINSPECTOR", "subinsp": "SUBINSPECTOR",
    "inspector": "INSPECTOR", "insp": "INSPECTOR",
    "comisionado mayor": "COMISIONADO MAYOR", "cdo mayor": "COMISIONADO MAYOR", "cdo. mayor": "COMISIONADO MAYOR", "com mayor": "COMISIONADO MAYOR",
    "comisionado general": "COMISIONADO GENERAL", "cdo general": "COMISIONADO GENERAL", "cdo. general": "COMISIONADO GENERAL", "cdo gral": "COMISIONADO GENERAL",
    "psa": "PSA", "aux": "AUXILIAR", "auxiliar": "AUXILIAR"
}

# --- ESTADOS ---
if 'analisis_listo' not in st.session_state: st.session_state.analisis_listo = False
if 'df_faltan' not in st.session_state: st.session_state.df_faltan = []
if 'df_sobran' not in st.session_state: st.session_state.df_sobran = pd.DataFrame()
if 'detective_candidates' not in st.session_state: st.session_state.detective_candidates = []
if 'total_parte' not in st.session_state: st.session_state.total_parte = 0
if 'total_lista' not in st.session_state: st.session_state.total_lista = 0
if 'checked_items' not in st.session_state: st.session_state.checked_items = set()
if 'confirmed_pairs' not in st.session_state: st.session_state.confirmed_pairs = {} 
if 'rejected_pairs' not in st.session_state: st.session_state.rejected_pairs = {}

# --- FUNCIONES DE LIMPIEZA ---
@st.cache_data
def normalizar_jerarquia(texto):
    if pd.isna(texto): return ""
    texto_limpio = str(texto).strip().lower()
    if texto_limpio in EQUIVALENCIAS: return EQUIVALENCIAS[texto_limpio]
    for key, value in EQUIVALENCIAS.items():
        if key in texto_limpio: return value
    return "" 

@st.cache_data
def limpiar_nombre(texto):
    if pd.isna(texto): return ""
    texto = str(texto)
    texto = re.sub(r'\([^)]*\)', '', texto)
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    texto = re.sub(r'[^a-zA-Z\s]', '', texto)
    return texto.strip().upper()

def abreviar_jerarquia(texto):
    if pd.isna(texto): return ""
    t = str(texto).upper()
    t = t.replace("OFICIAL", "OF")
    t = t.replace("AYUDANTE", "AYTE")
    t = t.replace("PRINCIPAL", "PPAL")
    return t

def leer_excel_inteligente(archivo_bytes, filename):
    try:
        xls = pd.ExcelFile(BytesIO(archivo_bytes))
        sheet_name = 'LISTA' if 'LISTA' in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        best_col_idx = -1; max_matches = 0
        for col_idx in range(len(df.columns)):
            col_data = df.iloc[:, col_idx].astype(str).str.lower()
            matches = col_data.apply(lambda x: any(k in x for k in EQUIVALENCIAS.keys())).sum()
            if matches > max_matches: max_matches = matches; best_col_idx = col_idx
        if best_col_idx != -1 and max_matches > 0 and best_col_idx + 1 < len(df.columns):
            subset = df.iloc[:, [best_col_idx, best_col_idx+1]].copy()
            subset.columns = ['Jerarquia', 'Nombre']
            return subset
        return None
    except: return None

def procesar_input(texto_input, archivo_input):
    df = None
    if archivo_input:
        file_bytes = archivo_input.getvalue()
        if archivo_input.name.endswith('csv'):
            try: df = pd.read_csv(archivo_input, header=None); df = df.iloc[:, :2]; df.columns = ['Jerarquia', 'Nombre']
            except: pass
        else:
            df = leer_excel_inteligente(file_bytes, archivo_input.name)
    elif texto_input:
        try:
            df = pd.read_csv(StringIO(texto_input), sep='\t', header=None, engine='python')
            if len(df.columns) < 2: df = pd.read_csv(StringIO(texto_input), sep=',', header=None, engine='python')
            df = df.iloc[:, :2]; df.columns = ['Jerarquia', 'Nombre']
        except: pass
    
    if df is not None and not df.empty:
        df['j_norm'] = df['Jerarquia'].apply(normalizar_jerarquia)
        df = df[df['j_norm'] != ""] 
        df['n_clean'] = df['Nombre'].apply(limpiar_nombre)
        df['unique_id'] = df['Nombre'] + "_" + df.index.astype(str)
        return df
    return None

# --- EXCEL CORE: HELPERS PARA CORRECCIÓN DE FORMATO Y COLOR ---

RED_FILL = PatternFill(fill_type="solid", fgColor="FFFF0000") # Rojo Intenso

def _merge_anchor(ws, row: int, col: int):
    """Si (row,col) cae dentro de un merge, devuelve la celda ancla (min_row,min_col)."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return row, col

def pintar_celda(ws, row: int, col: int, fill):
    r, c = _merge_anchor(ws, row, col)
    ws.cell(row=r, column=c).fill = fill

def pintar_cambio(ws, row: int, col_jerarquia: int, col_nombre: int, fill=RED_FILL):
    # pinta jerarquía + nombre (respetando merges)
    pintar_celda(ws, row, col_jerarquia, fill)
    pintar_celda(ws, row, col_nombre, fill)

def desplazar_merges_por_insercion(ws, fila_insercion: int, cantidad: int):
    old_ranges = [str(rng) for rng in ws.merged_cells.ranges]
    ws.merged_cells = MultiCellRange() 

    for r in old_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(r)
        if max_row < fila_insercion:
            new_r = r 
        elif min_row >= fila_insercion:
            new_r = f"{get_column_letter(min_col)}{min_row + cantidad}:{get_column_letter(max_col)}{max_row + cantidad}"
        else:
            new_r = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row + cantidad}"
        ws.merge_cells(new_r)

def snapshot_row_dims(ws, desde_fila: int):
    snap = {}
    for r, dim in list(ws.row_dimensions.items()):
        if r >= desde_fila:
            snap[r] = {
                "height": dim.height, "hidden": dim.hidden, "outlineLevel": dim.outlineLevel,
                "collapsed": dim.collapsed, "thickTop": dim.thickTop, "thickBot": dim.thickBot
            }
    return snap

def aplicar_row_dims_corridos(ws, snap, desde_fila: int, cantidad: int):
    for r in list(ws.row_dimensions.keys()):
        if r >= desde_fila: del ws.row_dimensions[r]
    for old_r in sorted(snap.keys()):
        new_r = old_r + cantidad
        dim = ws.row_dimensions[new_r]
        props = snap[old_r]
        for k, v in props.items(): 
            if hasattr(dim, k): setattr(dim, k, v)

# --- GENERADORES EXCEL ---

def borrar_sobrantes_excel(archivo_original, lista_nombres_borrar):
    # Función "Solo Borrar"
    try:
        wb = openpyxl.load_workbook(archivo_original)
        sheet_name = 'LISTA' if 'LISTA' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        col_jerarquia = -1; col_nombre = -1; max_matches = 0
        for col in range(1, 20):
            matches = 0
            for row in range(1, 50):
                val = str(ws.cell(row=row, column=col).value).lower()
                if any(k in val for k in EQUIVALENCIAS.keys()): matches += 1
            if matches > max_matches: max_matches = matches; col_jerarquia = col; col_nombre = col + 1 
        if col_jerarquia == -1: return None 
        
        nombres_a_borrar_limpios = set([limpiar_nombre(n) for n in lista_nombres_borrar])
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cell_nombre = row[col_nombre - 1]
            if not cell_nombre.value: continue
            val_nombre_limpio = limpiar_nombre(str(cell_nombre.value))
            if val_nombre_limpio in nombres_a_borrar_limpios:
                ws.cell(row=cell_nombre.row, column=col_jerarquia).value = None
                cell_nombre.value = None
                # 🔴 Marcar cambio en rojo al borrar
                pintar_cambio(ws, cell_nombre.row, col_jerarquia, col_nombre, RED_FILL)

        output = BytesIO(); wb.save(output); output.seek(0)
        return output
    except Exception as e:
        st.error(f"⚠️ Error al borrar: {e}")
        return None

def generar_excel_completo(archivo_original, lista_borrar, lista_agregar_dicts):
    # Función "Todo en Uno": Borra + Agrega
    try:
        wb = openpyxl.load_workbook(archivo_original)
        sheet_name = 'LISTA' if 'LISTA' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        
        col_jerarquia = -1; col_nombre = -1; max_matches = 0
        for col in range(1, 20):
            matches = 0
            for row in range(1, 50):
                val = str(ws.cell(row=row, column=col).value).lower()
                if any(k in val for k in EQUIVALENCIAS.keys()): matches += 1
            if matches > max_matches: max_matches = matches; col_jerarquia = col; col_nombre = col + 1 
        if col_jerarquia == -1: return None 

        # 1. BORRAR (Si hay lista de borrado)
        if lista_borrar:
            nombres_a_borrar_limpios = set([limpiar_nombre(n) for n in lista_borrar])
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                cell_nombre = row[col_nombre - 1]
                if not cell_nombre.value: continue
                val_nombre_limpio = limpiar_nombre(str(cell_nombre.value))
                if val_nombre_limpio in nombres_a_borrar_limpios:
                    ws.cell(row=cell_nombre.row, column=col_jerarquia).value = None
                    cell_nombre.value = None
                    # 🔴 Marcar cambio en rojo al borrar
                    pintar_cambio(ws, cell_nombre.row, col_jerarquia, col_nombre, RED_FILL)

        # 2. INSERTAR
        target_row = -1
        for row in range(1, ws.max_row + 1):
            row_values = [str(ws.cell(row=row, column=c).value).upper() for c in range(1, 10)]
            row_str = " ".join([v for v in row_values if v != 'None'])
            if "ARRIBO A2" in row_str or "ARRIBOS A2" in row_str:
                target_row = row
                break
        
        if target_row != -1 and len(lista_agregar_dicts) > 0:
            count = len(lista_agregar_dicts)
            snap_dims = snapshot_row_dims(ws, target_row)
            
            ws.insert_rows(target_row, amount=count)
            desplazar_merges_por_insercion(ws, target_row, count)
            aplicar_row_dims_corridos(ws, snap_dims, target_row, count)
            
            source_row_idx = target_row - 1
            model_dim = ws.row_dimensions[source_row_idx]
            for i in range(count):
                rd = ws.row_dimensions[target_row + i]
                if model_dim.height is not None:
                    rd.height = model_dim.height
            
            for i, persona in enumerate(lista_agregar_dicts):
                current_row = target_row + i
                for col in range(1, ws.max_column + 1):
                    source_cell = ws.cell(row=source_row_idx, column=col)
                    target_cell = ws.cell(row=current_row, column=col)
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = copy(source_cell.number_format)
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)
                
                jerarquia_corta = abreviar_jerarquia(str(persona['Jerarquia']))
                ws.cell(row=current_row, column=col_jerarquia).value = jerarquia_corta
                ws.cell(row=current_row, column=col_nombre).value = str(persona['Nombre']).upper()
                
                # 🔴 Marcar cambio en rojo al agregar (DESPUÉS DE COPIAR ESTILOS)
                pintar_cambio(ws, current_row, col_jerarquia, col_nombre, RED_FILL)

        output = BytesIO(); wb.save(output); output.seek(0)
        return output
    except Exception as e:
        st.error(f"❌ Error crítico generando el Excel: {e}")
        return None

# --- ANALISIS ---
def detecting_duplicados(df, nombre_origen):
    if df is None or df.empty: return
    duplicados = df[df.duplicated(subset=['n_clean'], keep=False)]
    if not duplicados.empty:
        nombres = [str(n) for n in duplicados['Nombre'].unique()]
        st.markdown(f'<div class="duplicate-alert">⚠️ <b>Duplicados en {nombre_origen}:</b> {", ".join(nombres[:3])}...</div>', unsafe_allow_html=True)

def calcular_analisis(df_p, df_l, umbral_det, umbral_auto):
    sobran = df_l.copy(); sobran['found'] = False
    faltan_temp = [] 
    for idx_p, row_p in df_p.iterrows():
        candidatos = sobran[sobran['j_norm'] == row_p['j_norm']]
        encontrado = False
        for idx_l, row_l in candidatos.iterrows():
            if row_l['found']: continue
            if fuzz.token_set_ratio(row_p['n_clean'], row_l['n_clean']) >= umbral_auto:
                encontrado = True; sobran.at[idx_l, 'found'] = True; break
        if not encontrado:
            for idx_l, row_l in sobran.iterrows():
                if row_l['found']: continue
                pair_id = f"{row_p['unique_id']}|{row_l['unique_id']}"
                if pair_id in st.session_state.confirmed_pairs:
                    encontrado = True; sobran.at[idx_l, 'found'] = True; break
        if not encontrado: faltan_temp.append(row_p)
    detective_matches = [] 
    df_sobran_reales = sobran[~sobran['found']]
    for f in faltan_temp:
        best_match = None; best_score = 0
        for idx_s, s in df_sobran_reales.iterrows():
            pair_id = f"{f['unique_id']}|{s['unique_id']}"
            if pair_id in st.session_state.rejected_pairs: continue 
            score = fuzz.token_sort_ratio(f['n_clean'], s['n_clean'])
            if score > umbral_det and score < umbral_auto: 
                if score > best_score: best_score = score; best_match = s
        if best_match is not None:
            detective_matches.append({'falta': f, 'sobra': best_match})
    return faltan_temp, df_sobran_reales, detective_matches

def ejecutar_analisis_completo(pf, lf):
    with st.spinner("Procesando..."):
        st.session_state.analisis_listo = False
        df_p = procesar_input(st.session_state.p_txt, pf)
        df_l = procesar_input(st.session_state.l_txt, lf)
        if df_p is not None and df_l is not None:
            detecting_duplicados(df_p, "PARTE")
            detecting_duplicados(df_l, "LISTA")
            u_auto = st.session_state.get('umbral_auto', 95)
            u_det = st.session_state.get('umbral_det', 65)
            faltan, sobran, detective = calcular_analisis(df_p, df_l, u_det, u_auto)
            st.session_state.df_faltan = faltan
            st.session_state.df_sobran = sobran
            st.session_state.detective_candidates = detective
            st.session_state.total_parte = len(df_p)
            st.session_state.total_lista = len(df_l)
            st.session_state.analisis_listo = True
        else:
            st.error("Error: Datos no válidos.")

# --- HISTORIAL & ACTIONS ---
def confirmar_match(f, s, pf, lf):
    st.session_state.confirmed_pairs[f"{f['unique_id']}|{s['unique_id']}"] = f"{f['Nombre']} ↔ {s['Nombre']}"
    ejecutar_analisis_completo(pf, lf)

def rechazar_match(f, s, pf, lf):
    st.session_state.rejected_pairs[f"{f['unique_id']}|{s['unique_id']}"] = f"{f['Nombre']} ≠ {s['Nombre']}"
    ejecutar_analisis_completo(pf, lf)

def deshacer_decision(pair_id, tipo, pf, lf):
    if tipo == 'confirmado': del st.session_state.confirmed_pairs[pair_id]
    elif tipo == 'rechazado': del st.session_state.rejected_pairs[pair_id]
    ejecutar_analisis_completo(pf, lf)

def limpiar_parte_callback(): st.session_state.p_txt = ""; st.session_state.p_key += 1; st.session_state.analisis_listo = False
def limpiar_lista_callback(): st.session_state.l_txt = ""; st.session_state.l_key += 1; st.session_state.analisis_listo = False

# --- UI INPUTS ---
if 'p_key' not in st.session_state: st.session_state.p_key = 0
if 'l_key' not in st.session_state: st.session_state.l_key = 0
if 'p_txt' not in st.session_state: st.session_state.p_txt = ""
if 'l_txt' not in st.session_state: st.session_state.l_txt = ""

col_c1, col_c2 = st.columns(2)
with col_c1:
    with st.container(border=True):
        h, b = st.columns([0.8, 0.2])
        h.markdown("### 📋 1. EL PARTE")
        b.button("Borrar", key="cl_p", on_click=limpiar_parte_callback, type="secondary")
        st.session_state.p_txt = st.text_area("P", height=120, key=f"p_txt_{st.session_state.p_key}", value=st.session_state.p_txt, label_visibility="collapsed", placeholder="Pegar Parte...")
        p_file = None 

with col_c2:
    with st.container(border=True):
        h, b = st.columns([0.8, 0.2])
        h.markdown("### 📝 2. LISTA GUARDIA")
        b.button("Borrar", key="cl_l", on_click=limpiar_lista_callback, type="secondary")
        l_file = st.file_uploader("L", type=["xlsx"], key=f"l_file_{st.session_state.l_key}", label_visibility="collapsed")
        with st.expander("O pegar texto"):
            st.session_state.l_txt = st.text_area("L", height=100, key=f"l_txt_{st.session_state.l_key}", value=st.session_state.l_txt, label_visibility="collapsed", placeholder="Pegar Lista...")

st.markdown("<br>", unsafe_allow_html=True)
if st.button("🔍 ANALIZAR AHORA", type="primary", use_container_width=True):
    ejecutar_analisis_completo(p_file, l_file)

# --- SIDEBAR ---
with st.sidebar:
    st.header("⚙️ Configuración")
    st.session_state.umbral_det = st.slider("Detective", 50, 90, 65)
    st.session_state.umbral_auto = st.slider("Automático", 80, 100, 95)
    st.divider()
    if st.session_state.confirmed_pairs:
        st.caption("Unidos")
        for pid, lbl in list(st.session_state.confirmed_pairs.items()):
            c1, c2 = st.columns([4,1])
            c1.caption(lbl)
            if c2.button("↩", key=f"dc_{pid}"): deshacer_decision(pid, 'confirmado', p_file, l_file); st.rerun()
    if st.session_state.rejected_pairs:
        st.caption("Separados")
        for pid, lbl in list(st.session_state.rejected_pairs.items()):
            c1, c2 = st.columns([4,1])
            c1.caption(lbl)
            if c2.button("↩", key=f"dr_{pid}"): deshacer_decision(pid, 'rechazado', p_file, l_file); st.rerun()

# --- RESULTADOS ---
if st.session_state.analisis_listo:
    st.divider()
    
    ids_conflict_f = [m['falta']['unique_id'] for m in st.session_state.detective_candidates]
    ids_conflict_s = [m['sobra']['unique_id'] for m in st.session_state.detective_candidates]
    final_verde = [f for f in st.session_state.df_faltan if f['unique_id'] not in ids_conflict_f]
    final_rojo = st.session_state.df_sobran[~st.session_state.df_sobran['unique_id'].isin(ids_conflict_s)]

    # --- ZONA DE DESCARGA ---
    if l_file is not None:
        st.markdown("### 📥 ACCIONES Y DESCARGAS")
        st.markdown('<div style="background-color: #111; padding: 15px; border-radius: 8px; border: 1px solid #444; margin-bottom: 25px;">', unsafe_allow_html=True)
        
        c1, c2 = st.columns(2)
        with c1:
            st.caption("Opción A: Solo Borrar Sobrantes")
            l_file.seek(0)
            xls_clean = borrar_sobrantes_excel(l_file, final_rojo['Nombre'].tolist())
            if xls_clean:
                st.download_button("🗑️ Solo Borrar", xls_clean, file_name=f"LIMPIO_{l_file.name}", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="secondary", use_container_width=True)
        
        with c2:
            st.caption("Opción B: Actualizar Todo (Borrar + Agregar)")
            l_file.seek(0)
            # PASAMOS LAS 3 COSAS: Archivo + Lista Borrar + Lista Agregar
            xls_full = generar_excel_completo(l_file, final_rojo['Nombre'].tolist(), final_verde)
            if xls_full:
                st.download_button("🔄 Actualizar Todo", xls_full, file_name=f"FINAL_{l_file.name}", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
        
        st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.detective_candidates:
        st.markdown("<br>", unsafe_allow_html=True)
        st.caption("🕵️ **CONFLICTOS DETECTADOS**")
        h_det = st.columns([3, 0.3, 3, 1, 1])
        h_det[0].caption("PARTE")
        h_det[2].caption("LISTA")
        
        for m in st.session_state.detective_candidates:
            f = m['falta']; s = m['sobra']
            cols = st.columns([3, 0.2, 3, 0.6, 0.6], vertical_alignment="center")
            with cols[0]: st.markdown(f'<div class="conflict-container"><div class="c-badge">{f["Jerarquia"]}</div><div class="c-name">{f["Nombre"]}</div></div>', unsafe_allow_html=True)
            with cols[1]: st.markdown('<div class="arrow-icon">↔</div>', unsafe_allow_html=True)
            with cols[2]: st.markdown(f'<div class="conflict-container"><div class="c-badge">{s["Jerarquia"]}</div><div class="c-name">{s["Nombre"]}</div></div>', unsafe_allow_html=True)
            with cols[3]: 
                if st.button("Misma", key=f"y_{f['unique_id']}", type="primary"): confirmar_match(f, s, p_file, l_file); st.rerun()
            with cols[4]: 
                if st.button("Distintos", key=f"n_{f['unique_id']}", type="secondary"): rechazar_match(f, s, p_file, l_file); st.rerun()
            st.markdown("<div style='margin-bottom: 6px;'></div>", unsafe_allow_html=True)

    st.markdown("<hr>", unsafe_allow_html=True)
    
    cr1, cr2 = st.columns(2)
    with cr1:
        st.markdown("### ✅ Falta Agregar")
        if not final_verde: st.success("Lista Completa.")
        else:
            h = st.columns([2, 4, 1.5])
            h[0].caption("JERARQUÍA")
            h[1].caption("NOMBRE")
            h[2].caption("ACCIÓN")
            st.markdown("<hr style='margin: 0;'>", unsafe_allow_html=True)
            for p in final_verde:
                checked = p['unique_id'] in st.session_state.checked_items
                r = st.columns([2, 4, 1.5], vertical_alignment="center")
                with r[0]: st.markdown(f'<div class="unified-text">{str(p["Jerarquia"]).upper()}</div>', unsafe_allow_html=True)
                with r[1]: 
                    if checked: st.markdown(f'<div class="unified-text name-ready">{str(p["Nombre"]).upper()}</div>', unsafe_allow_html=True)
                    else: st.code(str(p["Nombre"]).upper(), language="text") 
                with r[2]: 
                    def toggle(uid):
                        if uid in st.session_state.checked_items: st.session_state.checked_items.remove(uid)
                        else: st.session_state.checked_items.add(uid)
                    lbl = "↩" if checked else "Listo"
                    kind = "secondary" if checked else "primary"
                    st.button(lbl, key=f"b_{p['unique_id']}", type=kind, on_click=toggle, args=(p['unique_id'],))
                st.markdown("<hr style='margin: 0; opacity: 0.1;'>", unsafe_allow_html=True)

    with cr2:
        st.markdown("### ❌ Sobra")
        if final_rojo.empty: st.success("Limpio.")
        else:
            st.dataframe(final_rojo[['Jerarquia', 'Nombre']], hide_index=True, use_container_width=True, height=500)

    st.markdown("<br><br>", unsafe_allow_html=True)
    m1, m2, m3, m4, m5 = st.columns(5)
    with m1: st.metric("Parte", st.session_state.total_parte)
    with m2: st.metric("Lista", st.session_state.total_lista)
    with m3: st.metric("Faltan", len(final_verde))
    with m4: st.metric("Sobran", len(final_rojo))
    with m5: st.metric("En Duda", len(st.session_state.detective_candidates), delta_color="off")
